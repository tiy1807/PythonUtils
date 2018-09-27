from pdfminer.pdfparser import PDFParser
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfpage import PDFTextExtractionNotAllowed
from pdfminer.pdfinterp import PDFResourceManager
from pdfminer.pdfinterp import PDFPageInterpreter
from pdfminer.pdfdevice import PDFDevice
from pdfminer.converter import TextConverter
from cStringIO import StringIO
from pdfminer.layout import LAParams
import logging
from pdfrw import PdfReader, PdfWriter
from openpyxl import load_workbook

# The path where the PDF File is stored
FILE_PATH = "C:\\Users\\owner\\OneDrive\\SkyDocuments\\Scouting\\Operations\\"

# This list of characters are ignored when finding the agenda item
INVALID_CHARS = [":"," "]

# The name of the log file.
LOG_FILE = "pdfreader.log"
logger = logging.getLogger(__name__)

# Set to true to default the group option - makes it easier to test with.
DEFAULT_GROUP = True

# The possible groups that papers may be assigned.
GROUPS = ["Adults_at_risk",
          "Census",
          "PMG",
          "Intimate_care"
          "Grants",
          "Risk",
          "Safeguarding",
          "Uniform",
          "Awards",
          "Safety",
          "POR"]

class Paper():

    pages_readpdf = PdfReader(FILE_PATH + "Operationsv2.pdf").pages
    
    def __init__(self, pages, item_nu, raw_pages):
        self.item_nu = item_nu
        self.date = 1
        self.name = find_paper_name(raw_pages[0])
        self.authors = "author1"

        # Writes PDF file
        outdata = PdfWriter()
        for page in pages:
            outdata.addpage(Paper.pages_readpdf[page])
        outdata.write(FILE_PATH + "output\\%s.pdf" % item_nu)
        logger.info("Created file operations_%s, with pages: %s", item_nu, pages)
        
        # Allows the ability to skip groups
        if DEFAULT_GROUP:
            self.group = "default_group"
        else:
            self.group = raw_input("What group is item %s with name %s" % (item_nu, name))
    
    def get_attrs(self):
        return {"item_nu": self.item_nu, "date": self.date, "group": self.group, "authors": self.authors, "name": self.name}
        

def find_item_nu(string):
    
    if "Item: " in text:
        logger.debug("Agenda Item: found in page")
        itemnu =  ""
        for char in text.split("Item")[1]:
            if char == "\n":
                break
            elif char not in INVALID_CHARS:
                itemnu += char
        logger.debug("Found item number %s", itemnu)
    else:
        itemnu = "-1"
        logger.info("Item number not found")
    return itemnu 

def find_paper_name(string):
 
    if "Paper: " in text:
        logger.debug("Paper: found in page")
        name =  ""
        in_name = False
        for char in text.split("Paper: ")[1]:
            if in_name:
                if char == "\n":
                    break
                else:
                    itemnu += char
            else:
                if char == "\n":
                    in_name = True
        logger.debug("Found paper number %s", name)
    else:
        name = "-1"
        logger.info("Paper name not found")
    return name
    
if __name__ == "__main__":
    logging.basicConfig(filename=LOG_FILE, level=logging.DEBUG, filemode="w")
    file = file("C:\\Users\\owner\\OneDrive\\SkyDocuments\\Scouting\\Operations\\Operationsv2.pdf","rb")
    rsrcmgr = PDFResourceManager()
    
    #codec = 'utf-8'
    laparams = LAParams()

    old_item_nu = "-1"
    pages = []
    page_number = 1
    pagenos = set()
    
    #for pagenumber in range(30):
    #    pagenos.add(pagenumber)
    
    papers = []
    raw_pages = []

    for page in PDFPage.get_pages(file, pagenos):
        print page_number
        logger.debug("Reading page %s", page_number)
        retstr = StringIO()
        device = TextConverter(rsrcmgr, retstr, laparams=laparams)
        interpreter = PDFPageInterpreter(rsrcmgr, device)
        interpreter.process_page(page)
        text = retstr.getvalue()
        retstr.close()
        new_item_nu = find_item_nu(text)
        raw_pages.append(text)
        logger.debug("Old item nu: %s, New item nu: %s", old_item_nu, new_item_nu)
        
        if old_item_nu != new_item_nu and old_item_nu != "-1" and new_item_nu != "-1":
            papers.append(Paper(pages, old_item_nu, raw_pages))
            pages = []
        if old_item_nu != "-1" or new_item_nu != "-1":
            logger.debug("Adding page_number %s to agenda item %s", page_number, old_item_nu)
            pages.append(page_number-1)
        page_number += 1
        if new_item_nu != "-1":
            old_item_nu = new_item_nu
   
    logger.info("Creating excel workbook")
    wb = load_workbook(FILE_PATH + "Operations papers.xlsx")
    ws = wb.active
    for row_num in range(len(papers)):
        attrs = papers[row_num].get_attrs()
        for col_num in range(len(attrs)):
            cell = ws.cell(row=row_num+1, column=col_num+1, value=attrs[col_num])
    wb.save()
        

    
